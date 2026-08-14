"""
BACT seasonal scoring — Watershed 2025 BACT Analysis.xlsx rules.

Read-only helpers. Ratings reproduce the workbook IFS / COUNTIF / GEOMEAN logic.
These are program scores, not regulatory compliance statuses.

E. coli geomean == 126 is an uncovered IFS case in the workbook (neither <126 nor >126).
"""
from __future__ import annotations

import math
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

# Default 2025 BACT season window from weekly headers C2:Q2 in 2025 BACT Analysis.xlsx
DEFAULT_SEASON_START = date(2025, 5, 18)
DEFAULT_SEASON_END = date(2025, 8, 24)

RATINGS = ("Excellent", "Good", "Fair", "Poor", "NA")

SOURCE_NOTE = (
    "Ratings reproduce the scoring logic in Watershed's 2025 BACT Analysis workbook."
)


def _numeric_values(values: Iterable[Any]) -> List[float]:
    out: List[float] = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)):
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                continue
            out.append(float(v))
    return out


def excel_geomean(values: Iterable[Any]) -> Optional[float]:
    """Excel GEOMEAN of positive numeric values; None if empty or non-positive (IFERROR → blank)."""
    xs = [v for v in _numeric_values(values) if v > 0]
    if not xs:
        return None
    return math.exp(sum(math.log(x) for x in xs) / len(xs))


def excel_average(values: Iterable[Any]) -> Optional[float]:
    xs = _numeric_values(values)
    if not xs:
        return None
    return sum(xs) / len(xs)


def rate_e_coli(values: Iterable[Any]) -> Dict[str, Any]:
    """
    Workbook E. coli Score (column S):
      IFS(B="", "NA",
          AND(R=0, B<126), "Excellent",
          AND(R>=1, B<126), "Good",
          AND(OR(R=0,R=1,R=2), B>126), "Fair",
          AND(R>=3, B>126), "Poor")
    R = COUNTIFS(values, ">235"); B = GEOMEAN(values)
    Equality B=126 has no IFS branch (Excel #N/A).
    """
    xs = _numeric_values(values)
    if not xs:
        return {
            "parameter": "E. coli",
            "sample_count": 0,
            "geomean": None,
            "average": None,
            "exceedance_count": 0,
            "warning_count": None,
            "severe_count": None,
            "rating": "NA",
            "uncovered_equality": False,
        }
    gm = excel_geomean(xs)
    exceedance = sum(1 for v in xs if v > 235)
    rating: Optional[str]
    uncovered = False
    if gm is None:
        rating = "NA"
    elif exceedance == 0 and gm < 126:
        rating = "Excellent"
    elif exceedance >= 1 and gm < 126:
        rating = "Good"
    elif exceedance in (0, 1, 2) and gm > 126:
        rating = "Fair"
    elif exceedance >= 3 and gm > 126:
        rating = "Poor"
    else:
        rating = "#N/A"
        uncovered = True
    return {
        "parameter": "E. coli",
        "sample_count": len(xs),
        "geomean": gm,
        "average": None,
        "exceedance_count": exceedance,
        "warning_count": None,
        "severe_count": None,
        "rating": rating,
        "uncovered_equality": uncovered,
        "note": (
            "Geomean equals 126; workbook IFS has no branch for this equality."
            if uncovered
            else None
        ),
    }


def _rate_warn_severe_temp_style(
    parameter: str,
    values: Iterable[Any],
    warn_gt: float,
    severe_ge: float,
) -> Dict[str, Any]:
    """
    Temperature / Turbidity / Phosphate / Nitrate workbook pattern:
      IFS(B="", "NA", R=0, "Excellent", S=0, "Good",
          OR(S=1,S=2) or S=1/S=2 listed, "Fair", S>=3, "Poor")
    Temperature lists S=1 and S=2 as separate Fair branches; equivalent to OR.
    """
    xs = _numeric_values(values)
    if not xs:
        return {
            "parameter": parameter,
            "sample_count": 0,
            "geomean": None,
            "average": None,
            "exceedance_count": None,
            "warning_count": 0,
            "severe_count": 0,
            "rating": "NA",
            "uncovered_equality": False,
        }
    warning = sum(1 for v in xs if v > warn_gt)
    severe = sum(1 for v in xs if v >= severe_ge)
    if warning == 0:
        rating = "Excellent"
    elif severe == 0:
        rating = "Good"
    elif severe in (1, 2):
        rating = "Fair"
    else:
        rating = "Poor"
    return {
        "parameter": parameter,
        "sample_count": len(xs),
        "geomean": None,
        "average": excel_average(xs),
        "exceedance_count": None,
        "warning_count": warning,
        "severe_count": severe,
        "rating": rating,
        "uncovered_equality": False,
    }


def rate_temperature(values: Iterable[Any]) -> Dict[str, Any]:
    # COUNTIF >28 ; COUNTIF >=31
    return _rate_warn_severe_temp_style("Temperature", values, 28.0, 31.0)


def rate_turbidity(values: Iterable[Any]) -> Dict[str, Any]:
    # COUNTIF >15 ; COUNTIF >=50
    return _rate_warn_severe_temp_style("Turbidity", values, 15.0, 50.0)


def rate_phosphate(values: Iterable[Any]) -> Dict[str, Any]:
    # COUNTIF >0.05 ; COUNTIF >=0.1
    return _rate_warn_severe_temp_style("Phosphate", values, 0.05, 0.1)


def rate_nitrate_mg_l(values: Iterable[Any]) -> Dict[str, Any]:
    # Workbook uses mg/L; COUNTIF >1.5 ; COUNTIF >3.0 (strict > for severe)
    xs = _numeric_values(values)
    if not xs:
        return {
            "parameter": "Nitrate",
            "sample_count": 0,
            "geomean": None,
            "average": None,
            "exceedance_count": None,
            "warning_count": 0,
            "severe_count": 0,
            "rating": "NA",
            "uncovered_equality": False,
        }
    warning = sum(1 for v in xs if v > 1.5)
    severe = sum(1 for v in xs if v > 3.0)
    if warning == 0:
        rating = "Excellent"
    elif severe == 0:
        rating = "Good"
    elif severe in (1, 2):
        rating = "Fair"
    else:
        rating = "Poor"
    return {
        "parameter": "Nitrate",
        "sample_count": len(xs),
        "geomean": None,
        "average": excel_average(xs),
        "exceedance_count": None,
        "warning_count": warning,
        "severe_count": severe,
        "rating": rating,
        "uncovered_equality": False,
    }


def rate_chloride(values: Iterable[Any]) -> Dict[str, Any]:
    """
    Chloride Score:
      IFS(B="", "NA", R=0, "Excellent", OR(S=0,S=1), "Good", S=2, "Fair", S>=3, "Poor")
    R = COUNTIF >100; S = COUNTIF >=230
    """
    xs = _numeric_values(values)
    if not xs:
        return {
            "parameter": "Chloride",
            "sample_count": 0,
            "geomean": None,
            "average": None,
            "exceedance_count": None,
            "warning_count": 0,
            "severe_count": 0,
            "rating": "NA",
            "uncovered_equality": False,
        }
    warning = sum(1 for v in xs if v > 100)
    severe = sum(1 for v in xs if v >= 230)
    if warning == 0:
        rating = "Excellent"
    elif severe in (0, 1):
        rating = "Good"
    elif severe == 2:
        rating = "Fair"
    else:
        rating = "Poor"
    return {
        "parameter": "Chloride",
        "sample_count": len(xs),
        "geomean": None,
        "average": excel_average(xs),
        "exceedance_count": None,
        "warning_count": warning,
        "severe_count": severe,
        "rating": rating,
        "uncovered_equality": False,
    }


def score_parameter_set(
    e_coli: Sequence[Any] = (),
    temperature_c: Sequence[Any] = (),
    turbidity_ntu: Sequence[Any] = (),
    chloride_mg_l: Sequence[Any] = (),
    phosphate_mg_l: Sequence[Any] = (),
    nitrate_mg_l: Sequence[Any] = (),
) -> List[Dict[str, Any]]:
    return [
        rate_e_coli(e_coli),
        rate_temperature(temperature_c),
        rate_turbidity(turbidity_ntu),
        rate_chloride(chloride_mg_l),
        rate_phosphate(phosphate_mg_l),
        rate_nitrate_mg_l(nitrate_mg_l),
    ]


def nitrate_ug_l_to_mg_l(ug_l: Any) -> Optional[float]:
    if ug_l is None:
        return None
    try:
        return float(ug_l) / 1000.0
    except (TypeError, ValueError):
        return None


def score_site_from_series(
    site_code: str,
    *,
    waterbody: Optional[str] = None,
    e_coli: Sequence[Any] = (),
    temperature_c: Sequence[Any] = (),
    turbidity_ntu: Sequence[Any] = (),
    chloride_mg_l: Sequence[Any] = (),
    phosphate_mg_l: Sequence[Any] = (),
    nitrate_mg_l: Sequence[Any] = (),
    sample_dates: Optional[Sequence[Any]] = None,
    input_source: str = "explicit series",
) -> Dict[str, Any]:
    """Score one site from an explicit measurement series (workbook or other)."""
    params = score_parameter_set(
        e_coli=e_coli,
        temperature_c=temperature_c,
        turbidity_ntu=turbidity_ntu,
        chloride_mg_l=chloride_mg_l,
        phosphate_mg_l=phosphate_mg_l,
        nitrate_mg_l=nitrate_mg_l,
    )
    dates = [str(d)[:10] for d in (sample_dates or []) if d is not None]
    return {
        "site_code": site_code,
        "waterbody": waterbody,
        "found": True,
        "date_start": dates[0] if dates else None,
        "date_end": dates[-1] if dates else None,
        "last_sample_date": dates[-1] if dates else None,
        "visit_count": max(
            len(_numeric_values(e_coli)),
            len(_numeric_values(temperature_c)),
            len(_numeric_values(turbidity_ntu)),
            len(_numeric_values(chloride_mg_l)),
            len(_numeric_values(phosphate_mg_l)),
            len(_numeric_values(nitrate_mg_l)),
            0,
        ),
        "parameters": params,
        "input_source": input_source,
        "source_note": SOURCE_NOTE,
    }


def _weekly_numeric_row(ws, row: int, first_col: int = 3, last_col: int = 17) -> List[Any]:
    """Read Analysis sheet weekly value cells C:Q (columns 3–17)."""
    return [ws.cell(row, c).value for c in range(first_col, last_col + 1)]


def score_bact_analysis_workbook(path) -> Dict[str, Any]:
    """
    Read-only seasonal scores from a saved 2025 BACT Analysis.xlsx.

    Uses cached weekly values on each parameter sheet (the same cells the workbook
    scores with GEOMEAN/COUNTIF/IFS). Does not query PostgreSQL.

    Requires a workbook saved with calculated values (Excel data cache).
    """
    from openpyxl import load_workbook

    path = Path(path) if not isinstance(path, Path) else path
    wb = load_workbook(path, data_only=True)
    required = ["Temperature", "E. coli", "Turbidity", "Chloride", "Phosphate", "Nitrate"]
    missing = [n for n in required if n not in wb.sheetnames]
    if missing:
        raise ValueError(
            "This file does not look like a 2025 BACT Analysis workbook "
            f"(missing sheets: {', '.join(missing)})."
        )

    temp_ws = wb["Temperature"]
    dates = []
    for c in range(3, 18):
        d = temp_ws.cell(2, c).value
        if d is not None:
            dates.append(d.date() if hasattr(d, "date") else d)

    # Optional waterbody from Dashboard_Values
    waterbody_by_site: Dict[str, str] = {}
    if "Dashboard_Values" in wb.sheetnames:
        dv = wb["Dashboard_Values"]
        for r in range(2, dv.max_row + 1):
            site = dv.cell(r, 1).value
            wbod = dv.cell(r, 2).value
            if site:
                waterbody_by_site[str(site)] = str(wbod) if wbod else None

    sheets = {
        "temperature_c": wb["Temperature"],
        "e_coli": wb["E. coli"],
        "turbidity_ntu": wb["Turbidity"],
        "chloride_mg_l": wb["Chloride"],
        "phosphate_mg_l": wb["Phosphate"],
        "nitrate_mg_l": wb["Nitrate"],
    }

    site_results: List[Dict[str, Any]] = []
    for r in range(3, temp_ws.max_row + 1):
        site = temp_ws.cell(r, 1).value
        if not site or str(site).strip() in ("", "AA1"):  # skip test stub if empty series only
            continue
        site = str(site).strip()
        series = {
            key: _weekly_numeric_row(ws, r)
            for key, ws in sheets.items()
        }
        # Skip sites with no numeric values on any parameter
        if not any(_numeric_values(v) for v in series.values()):
            continue
        site_results.append(
            score_site_from_series(
                site,
                waterbody=waterbody_by_site.get(site),
                e_coli=series["e_coli"],
                temperature_c=series["temperature_c"],
                turbidity_ntu=series["turbidity_ntu"],
                chloride_mg_l=series["chloride_mg_l"],
                phosphate_mg_l=series["phosphate_mg_l"],
                nitrate_mg_l=series["nitrate_mg_l"],
                sample_dates=dates,
                input_source=(
                    "2025 BACT Analysis.xlsx weekly parameter sheets "
                    "(SURVEY123-linked season grid)"
                ),
            )
        )

    return {
        "site_results": site_results,
        "season_dates": [str(d)[:10] for d in dates],
        "input_source": "2025 BACT Analysis.xlsx",
        "source_note": SOURCE_NOTE,
        "authority_note": (
            "Scores use the Analysis workbook’s weekly selected values. "
            "Those cells XLOOKUP into BACT and HAB SURVEY123 (field temp; "
            "IDEXX E. coli; Gallery chloride/phosphate/nitrate; TURBIDITY NTU). "
            "They are not derived from arbitrary PostgreSQL chemical packages."
        ),
    }


def dashboard_values_row(site_result: Dict[str, Any]) -> Dict[str, Any]:
    """
    AGO-preparation style summary fields from Dashboard_Values that map unambiguously.
    Omits per-week Temp_*/Ecoli_* columns (season-grid specific).
    """
    by_name = {p["parameter"]: p for p in site_result.get("parameters") or []}

    def rating(name):
        p = by_name.get(name) or {}
        return p.get("rating")

    def avg(name):
        p = by_name.get(name) or {}
        return p.get("average")

    ecoli = by_name.get("E. coli") or {}
    return {
        "Site": site_result.get("site_code"),
        "Waterbody": site_result.get("waterbody"),
        "LastSampleDate": site_result.get("last_sample_date"),
        "SampleNumber": site_result.get("visit_count"),
        "Temperature_Score": rating("Temperature"),
        "AvgTemp": avg("Temperature"),
        "E__Coli_Score": rating("E. coli"),
        "EcoliGeoMean": ecoli.get("geomean"),
        "Turbidity_Score": rating("Turbidity"),
        "AvgTurbidity": avg("Turbidity"),
        "Chloride_Score": rating("Chloride"),
        "AvgChloride": avg("Chloride"),
        "Phosphate_Score": rating("Phosphate"),
        "AvgPhosphate": avg("Phosphate"),
        "Nitrate_Score": rating("Nitrate"),
        "AvgNitrate": avg("Nitrate"),
    }
