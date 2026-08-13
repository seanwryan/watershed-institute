#!/usr/bin/env python3
"""
Migrate CAT Meter Tracking v.1.xlsx into equipment, sensor, session, meter_maintenance, meter_testing.
Sheets: Assignments (inventory), Sensors (replacement log), Tracking (failure/quarterly),
2024/2025/2026 Testing (wide quarterly matrices → session + meter_testing).
Expects: STREAMWATCH_DATA_DIR or EQUIPMENT_FILE.
"""
import pandas as pd
from pathlib import Path
import sys
import re
from datetime import datetime, date

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from etl.config import EQUIPMENT_FILE
from etl.db import get_conn, ensure_lookup

# Real CAT multiparameter meters only (Assignments also lists LaMotte users under Meter ID).
EQUIPMENT_CODE_RE = re.compile(r"^TWI\d{3}$", re.IGNORECASE)
ROUND_RE = re.compile(r"^Round\s+(\d+)$", re.IGNORECASE)


def _str(v):
    if v is None or pd.isna(v):
        return None
    return str(v).strip() or None


def _serial(v):
    """Normalize equipment serial numbers without turning text SNs through float (keeps leading zeros)."""
    if v is None or pd.isna(v):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip() or None
    s = str(v).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+\.0", s):
        return s[:-2]
    return s


def _int(v):
    if v is None or pd.isna(v):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _float(v):
    if v is None or pd.isna(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _date(v):
    if v is None or pd.isna(v):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if hasattr(v, "date") and callable(v.date):
        try:
            return v.date()
        except Exception:
            pass
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None


def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _norm_label(v):
    if not isinstance(v, str):
        return None
    return re.sub(r"\s+", " ", v.strip().lower()).rstrip(".")


def _normalize_meter_code(v):
    """Strip trailing * markers; accept only TWI + exactly 3 digits. Returns uppercase code or None."""
    if not isinstance(v, str):
        return None
    s = re.sub(r"\*+$", "", v.strip()).strip()
    if EQUIPMENT_CODE_RE.fullmatch(s):
        return s.upper()
    return None


def _classify_meas_columns(ws, header_row, meter_col):
    """Map Meter-header labels to measured/diff columns for pH, DO ppm, and primary Cond/EC."""
    labels = {}
    blank_run = 0
    for cc in range(meter_col + 1, min(meter_col + 14, ws.max_column + 1)):
        raw = ws.cell(header_row, cc).value
        if raw is None or (isinstance(raw, str) and not str(raw).strip()):
            blank_run += 1
            if blank_run >= 2:
                break
            continue
        blank_run = 0
        if isinstance(raw, str) and raw.strip() == "Meter":
            break
        lab = _norm_label(raw)
        if lab:
            labels[lab] = cc

    return {
        "pH": {
            "meas": labels.get("ph"),
            "diff": labels.get("ph diff"),
            "std_keys": ("ph",),
        },
        "DO": {
            "meas": labels.get("do ppm"),
            "diff": None,
            "std_keys": ("do ppm",),
        },
        # Prefer Cond.; else Cond. 1. Do not use Cond. 2 in this pass.
        "EC": {
            "meas": labels.get("cond") or labels.get("cond 1"),
            "diff": labels.get("cond diff") or labels.get("cond 1 diff"),
            "std_keys": ("cond", "cond 1"),
        },
    }


def _parse_testing_sheet(ws, year):
    """
    Parse a wide quarterly Testing sheet into dated round blocks with measurements.
    Skips blocks without a confident test date. Does not invent pass/fail or staff.
    Returns (blocks, stats) where each block is:
      {year, round_number, test_date, meter_col, measurements: [dicts]}
    """
    stats = {
        "blocks_ok": 0,
        "blocks_skipped_no_date": 0,
        "blocks_skipped_no_meter_header": 0,
        "invalid_meter_cells": 0,
    }
    blocks = []

    round_cells = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                m = ROUND_RE.fullmatch(v.strip())
                if m:
                    round_cells.append((r, c, int(m.group(1))))

    for round_row, meter_col, round_number in round_cells:
        test_date = None
        for r in range(round_row, min(round_row + 4, ws.max_row + 1)):
            for c in range(meter_col, min(meter_col + 3, ws.max_column + 1)):
                if ws.cell(r, c).value == "Test Date:":
                    test_date = _date(ws.cell(r, c + 1).value)
                    break
            if test_date:
                break
        if not test_date:
            stats["blocks_skipped_no_date"] += 1
            continue

        meter_header_row = None
        for r in range(round_row + 1, min(round_row + 12, ws.max_row + 1)):
            if ws.cell(r, meter_col).value == "Meter":
                meter_header_row = r
                break
        if not meter_header_row:
            stats["blocks_skipped_no_meter_header"] += 1
            continue

        std_by_label = {}
        for r in range(round_row, meter_header_row):
            if ws.cell(r, meter_col).value == "Standards:":
                for c in range(meter_col + 1, min(meter_col + 14, ws.max_column + 1)):
                    lab = _norm_label(ws.cell(r, c).value)
                    if lab:
                        std_by_label[lab] = ws.cell(r + 1, c).value
                break

        colmap = _classify_meas_columns(ws, meter_header_row, meter_col)
        measurements = []

        for r in range(meter_header_row + 1, min(meter_header_row + 40, ws.max_row + 1)):
            raw = ws.cell(r, meter_col).value
            if isinstance(raw, str) and raw.strip().upper().startswith("NOTES"):
                break
            if raw is None:
                continue
            meter_code = _normalize_meter_code(raw)
            if meter_code is None:
                if isinstance(raw, str) and raw.strip():
                    stats["invalid_meter_cells"] += 1
                continue

            for parameter_type, cfg in (
                ("pH", colmap["pH"]),
                ("DO", colmap["DO"]),
                ("EC", colmap["EC"]),
            ):
                meas_col = cfg["meas"]
                if not meas_col:
                    continue
                measured = ws.cell(r, meas_col).value
                # Require a real numeric measured value (skips blanks, text, and template-only diffs).
                if not _is_number(measured):
                    continue

                reference_value = None
                for sk in cfg["std_keys"]:
                    if sk in std_by_label and _is_number(std_by_label[sk]):
                        reference_value = float(std_by_label[sk])
                        break

                difference = None
                if cfg["diff"]:
                    dv = ws.cell(r, cfg["diff"]).value
                    if _is_number(dv):
                        difference = float(dv)

                measurements.append({
                    "meter_code": meter_code,
                    "parameter_type": parameter_type,
                    "reference_value": reference_value,
                    "measured_value": float(measured),
                    "difference": difference,
                })

        stats["blocks_ok"] += 1
        blocks.append({
            "year": year,
            "round_number": round_number,
            "test_date": test_date,
            "meter_col": meter_col,
            "measurements": measurements,
        })

    return blocks, stats


def run():
    if not EQUIPMENT_FILE.exists():
        print(f"Equipment file not found: {EQUIPMENT_FILE}. Set STREAMWATCH_DATA_DIR or EQUIPMENT_FILE.")
        sys.exit(1)

    xl = pd.ExcelFile(EQUIPMENT_FILE)
    assignments_df = pd.read_excel(EQUIPMENT_FILE, sheet_name="Assignments") if "Assignments" in xl.sheet_names else pd.DataFrame()
    sensors_df = pd.read_excel(EQUIPMENT_FILE, sheet_name="Sensors") if "Sensors" in xl.sheet_names else pd.DataFrame()
    tracking_df = pd.read_excel(EQUIPMENT_FILE, sheet_name="Tracking") if "Tracking" in xl.sheet_names else pd.DataFrame()

    with get_conn() as conn:
        cur = conn.cursor()
        session_type_id = ensure_lookup(conn, "lst_session_type", "session_type_id", "name", "Quarterly maintenance")

        equipment_code_to_id = {}
        accepted_equipment = 0
        skipped_non_equipment = 0

        for _, row in assignments_df.iterrows():
            meter_id = _str(row.get("Meter ID") or row.get("MeterID") or row.get("MeterId"))
            if not meter_id:
                continue
            if not EQUIPMENT_CODE_RE.fullmatch(meter_id):
                skipped_non_equipment += 1
                continue
            meter_id = meter_id.upper()
            sn = _serial(row.get("Serial number") or row.get("SN") or row.get("Serial Number"))
            status = "Active"
            if _str(row.get("Retired")) or str(row.get("Inactive", "")).strip().lower() in ("1", "true", "yes", "x"):
                status = "Retired"
            cur.execute("""
                INSERT INTO equipment (equipment_code, equipment_type, serial_number, status)
                VALUES (%s, 'Multiparameter meter', %s, %s::equipment_status_enum)
                ON CONFLICT (equipment_code) DO UPDATE SET serial_number = EXCLUDED.serial_number, status = EXCLUDED.status, updated_at = NOW()
                RETURNING equipment_id
                """, (meter_id, sn, status))
            eq_id = cur.fetchone()[0]
            equipment_code_to_id[meter_id] = eq_id
            accepted_equipment += 1

        for _, row in sensors_df.iterrows():
            meter_id = _str(row.get("Meter ID") or row.get("MeterID"))
            if meter_id:
                meter_id = meter_id.upper()
            if not meter_id or meter_id not in equipment_code_to_id:
                continue
            eq_id = equipment_code_to_id[meter_id]
            for param, sensor_type in (("DO", "DO"), ("pH", "pH"), ("EC", "EC")):
                col = _str(row.get(f"Date last changed {param}") or row.get(f"{param}") or row.get("DO sensor") or row.get("pH sensor") or row.get("EC sensor"))
                date_installed = _date(col) if col else None
                if date_installed or param in str(row.keys()):
                    cur.execute("""
                        INSERT INTO sensor (equipment_id, sensor_type, date_installed) VALUES (%s, %s::sensor_type_enum, %s)
                        """, (eq_id, param, date_installed))

        testing_sheets_found = []
        sessions_created = 0
        meter_testing_by_year = {"2024": 0, "2025": 0, "2026": 0}
        blocks_skipped_no_date = 0
        blocks_skipped_no_meter_header = 0
        invalid_meter_cells = 0
        skipped_unknown_equipment = 0

        xlsx = load_workbook(EQUIPMENT_FILE, data_only=True)
        for year in ("2024", "2025", "2026"):
            sheet_name = f"{year} Testing"
            if sheet_name not in xlsx.sheetnames:
                continue
            testing_sheets_found.append(sheet_name)
            blocks, stats = _parse_testing_sheet(xlsx[sheet_name], year)
            blocks_skipped_no_date += stats["blocks_skipped_no_date"]
            blocks_skipped_no_meter_header += stats["blocks_skipped_no_meter_header"]
            invalid_meter_cells += stats["invalid_meter_cells"]

            for block in blocks:
                session_id = None
                for meas in block["measurements"]:
                    eq_id = equipment_code_to_id.get(meas["meter_code"])
                    if not eq_id:
                        skipped_unknown_equipment += 1
                        continue
                    if session_id is None:
                        summary = (
                            f"CAT meter testing {block['year']} "
                            f"Round {block['round_number']} ({block['test_date'].isoformat()})"
                        )
                        cur.execute("""
                            INSERT INTO session (session_type_id, date_start, date_end, summary)
                            VALUES (%s, %s, %s, %s) RETURNING session_id
                            """, (session_type_id, block["test_date"], block["test_date"], summary))
                        session_id = cur.fetchone()[0]
                        sessions_created += 1
                    cur.execute("""
                        INSERT INTO meter_testing (
                            session_id, equipment_id, parameter_type, round_number,
                            reference_value, measured_value, difference
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (
                            session_id,
                            eq_id,
                            meas["parameter_type"],
                            block["round_number"],
                            meas["reference_value"],
                            meas["measured_value"],
                            meas["difference"],
                        ))
                    meter_testing_by_year[year] += 1

    print(
        f"Equipment migration done. "
        f"Accepted equipment: {accepted_equipment}; "
        f"skipped non-equipment Meter IDs: {skipped_non_equipment}."
    )
    print(
        f"Meter testing: sheets found={testing_sheets_found or 'none'}; "
        f"sessions created={sessions_created}; "
        f"meter_testing rows by year={dict(meter_testing_by_year)}; "
        f"total meter_testing={sum(meter_testing_by_year.values())}; "
        f"blocks skipped (no date)={blocks_skipped_no_date}; "
        f"blocks skipped (no Meter header)={blocks_skipped_no_meter_header}; "
        f"invalid meter cells={invalid_meter_cells}; "
        f"unknown equipment skips={skipped_unknown_equipment}."
    )


if __name__ == "__main__":
    run()
